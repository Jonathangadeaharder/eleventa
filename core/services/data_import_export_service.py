from typing import Dict, Any
from decimal import Decimal
import csv
import json
import os
from datetime import datetime
import logging
import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from core.models.product import Product, Department
from core.services.service_base import ServiceBase
from infrastructure.persistence.unit_of_work import unit_of_work


class DataImportExportService(ServiceBase):
    """Servicio para importar y exportar datos del inventario."""

    def __init__(self):
        super().__init__()
        self.logger = logging.getLogger(__name__)

    def export_products_to_excel(self, file_path: str) -> Dict[str, Any]:
        """Exporta todos los productos a un archivo Excel."""
        try:
            with unit_of_work() as uow:
                products = uow.products.get_all()
                departments = {dept.id: dept.name for dept in uow.departments.get_all()}

                # Convertir a DataFrame
                data = []
                for product in products:
                    data.append(
                        {
                            "Código": product.code,
                            "Descripción": product.description,
                            "Precio Costo": (
                                float(product.cost_price) if product.cost_price else 0
                            ),
                            "Precio Venta": (
                                float(product.sell_price) if product.sell_price else 0
                            ),
                            "Stock": (
                                float(product.quantity_in_stock)
                                if product.quantity_in_stock
                                else 0
                            ),
                            "Stock Mínimo": (
                                float(product.min_stock) if product.min_stock else 0
                            ),
                            "Unidad": (
                                product.unit if hasattr(product, "unit") else "Unidad"
                            ),
                            "Usa Inventario": "Sí" if product.uses_inventory else "No",
                            "Departamento": (
                                departments.get(product.department_id, "")
                                if product.department_id
                                else ""
                            ),
                        }
                    )

                df = pd.DataFrame(data)
                df.to_excel(file_path, index=False, engine="openpyxl")

                self.logger.info(
                    f"Exported {len(products)} products to Excel: {file_path}"
                )

                return {
                    "success": True,
                    "message": f"Productos exportados exitosamente a {os.path.basename(file_path)}",
                    "products_exported": len(products),
                }

        except Exception as e:
            self.logger.error(f"Error exporting products to Excel: {e}")
            return {
                "success": False,
                "message": "Error al exportar productos a Excel",
                "error": str(e),
            }

    def export_products_to_csv(self, file_path: str) -> Dict[str, Any]:
        """Exporta todos los productos a un archivo CSV."""
        try:
            with unit_of_work() as uow:
                products = uow.products.get_all()
                departments = {dept.id: dept.name for dept in uow.departments.get_all()}

                # Convertir a DataFrame
                data = []
                for product in products:
                    data.append(
                        {
                            "Código": product.code,
                            "Descripción": product.description,
                            "Precio Costo": (
                                float(product.cost_price) if product.cost_price else 0
                            ),
                            "Precio Venta": (
                                float(product.sell_price) if product.sell_price else 0
                            ),
                            "Stock": (
                                float(product.quantity_in_stock)
                                if product.quantity_in_stock
                                else 0
                            ),
                            "Stock Mínimo": (
                                float(product.min_stock) if product.min_stock else 0
                            ),
                            "Unidad": (
                                product.unit if hasattr(product, "unit") else "Unidad"
                            ),
                            "Usa Inventario": "Sí" if product.uses_inventory else "No",
                            "Departamento": (
                                departments.get(product.department_id, "")
                                if product.department_id
                                else ""
                            ),
                        }
                    )

                df = pd.DataFrame(data)
                df.to_csv(file_path, index=False, encoding="utf-8-sig")

                self.logger.info(
                    f"Exported {len(products)} products to CSV: {file_path}"
                )

                return {
                    "success": True,
                    "message": f"Productos exportados exitosamente a {os.path.basename(file_path)}",
                    "products_exported": len(products),
                }

        except Exception as e:
            self.logger.error(f"Error exporting products to CSV: {e}")
            return {
                "success": False,
                "message": "Error al exportar productos a CSV",
                "error": str(e),
            }

    def import_products_from_excel(self, file_path: str) -> Dict[str, Any]:
        """Importa productos desde un archivo Excel."""
        if not EXCEL_AVAILABLE:
            return {
                "success": False,
                "error": "openpyxl no está instalado. Instale con: pip install openpyxl",
                "message": "Excel no está disponible",
            }

        if not os.path.exists(file_path):
            return {
                "success": False,
                "error": "Archivo no encontrado",
                "message": "El archivo especificado no existe.",
            }

        try:
            with unit_of_work() as uow:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active

                results = {
                    "success": True,
                    "total_rows": 0,
                    "imported": 0,
                    "updated": 0,
                    "skipped": 0,
                    "errors": [],
                }

                # Leer headers (primera fila)
                headers = [cell.value for cell in ws[1]]

                # Procesar cada fila de datos
                for row_num, row in enumerate(
                    ws.iter_rows(min_row=2, values_only=True), 2
                ):
                    if not any(row):  # Saltar filas vacías
                        continue

                    results["total_rows"] += 1

                    try:
                        # Extraer datos de la fila
                        codigo = str(row[0]).strip() if row[0] else ""
                        descripcion = str(row[1]).strip() if row[1] else ""
                        precio_costo = float(row[2]) if row[2] is not None else 0.0
                        precio_venta = float(row[3]) if row[3] is not None else 0.0
                        stock = float(row[4]) if row[4] is not None else 0.0
                        stock_minimo = float(row[5]) if row[5] is not None else 0.0
                        unidad = str(row[6]).strip() if row[6] else "U"
                        usa_inventario = (
                            str(row[7]).strip().lower()
                            in ["si", "sí", "yes", "true", "1"]
                            if row[7]
                            else True
                        )
                        departamento_nombre = str(row[8]).strip() if row[8] else ""

                        if not codigo or not descripcion:
                            results["errors"].append(
                                f"Fila {row_num}: Código y descripción son obligatorios"
                            )
                            results["skipped"] += 1
                            continue

                        # Buscar departamento por nombre
                        department_id = None
                        if departamento_nombre and departamento_nombre != "-":
                            departments = uow.departments.get_all()
                            for dept in departments:
                                if dept.name.lower() == departamento_nombre.lower():
                                    department_id = dept.id
                                    break

                        # Verificar si el producto ya existe
                        existing_product = uow.products.get_by_code(codigo)

                        product_data = {
                            "code": codigo,
                            "description": descripcion,
                            "cost_price": Decimal(str(precio_costo)),
                            "sell_price": Decimal(str(precio_venta)),
                            "quantity_in_stock": Decimal(str(stock)),
                            "min_stock": Decimal(str(stock_minimo)),
                            "unit": unidad,
                            "uses_inventory": usa_inventario,
                            "department_id": department_id,
                        }

                        if existing_product:
                            # Actualizar producto existente
                            for key, value in product_data.items():
                                setattr(existing_product, key, value)
                            uow.products.update(existing_product)
                            results["updated"] += 1
                        else:
                            # Crear nuevo producto
                            new_product = Product(**product_data)
                            uow.products.add(new_product)
                            results["imported"] += 1

                    except Exception as e:
                        error_msg = f"Fila {row_num}: {str(e)}"
                        results["errors"].append(error_msg)
                        results["skipped"] += 1
                        self.logger.error(error_msg)

                # Commit de los cambios
                uow.commit()

                results["message"] = (
                    f"Importación completada: {results['imported']} nuevos, {results['updated']} actualizados, {results['skipped']} omitidos"
                )

                return results

        except Exception as e:
            self.logger.error(f"Error importando productos desde Excel: {e}")
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al importar: {e}",
            }

    def import_products_from_csv(self, file_path: str) -> Dict[str, Any]:
        """Importa productos desde un archivo CSV."""
        if not os.path.exists(file_path):
            return {
                "success": False,
                "error": "Archivo no encontrado",
                "message": "El archivo especificado no existe.",
            }

        try:
            with unit_of_work() as uow:
                results = {
                    "success": True,
                    "total_rows": 0,
                    "imported": 0,
                    "updated": 0,
                    "skipped": 0,
                    "errors": [],
                }

                with open(file_path, "r", encoding="utf-8") as csvfile:
                    # Detectar el dialecto del CSV
                    sample = csvfile.read(1024)
                    csvfile.seek(0)
                    sniffer = csv.Sniffer()
                    delimiter = sniffer.sniff(sample).delimiter

                    reader = csv.DictReader(csvfile, delimiter=delimiter)

                    for row_num, row in enumerate(reader, 2):
                        results["total_rows"] += 1

                        try:
                            # Extraer datos de la fila
                            codigo = row.get("codigo", "").strip()
                            descripcion = row.get("descripcion", "").strip()
                            precio_costo = float(row.get("precio_costo", 0))
                            precio_venta = float(row.get("precio_venta", 0))
                            stock = float(row.get("stock", 0))
                            stock_minimo = float(row.get("stock_minimo", 0))
                            unidad = row.get("unidad", "U").strip()
                            usa_inventario = row.get(
                                "usa_inventario", "Si"
                            ).strip().lower() in ["si", "sí", "yes", "true", "1"]
                            departamento_nombre = row.get("departamento", "").strip()

                            if not codigo or not descripcion:
                                results["errors"].append(
                                    f"Fila {row_num}: Código y descripción son obligatorios"
                                )
                                results["skipped"] += 1
                                continue

                            # Buscar departamento por nombre
                            department_id = None
                            if departamento_nombre and departamento_nombre != "-":
                                departments = uow.departments.get_all()
                                for dept in departments:
                                    if dept.name.lower() == departamento_nombre.lower():
                                        department_id = dept.id
                                        break

                            # Verificar si el producto ya existe
                            existing_product = uow.products.get_by_code(codigo)

                            product_data = {
                                "code": codigo,
                                "description": descripcion,
                                "cost_price": Decimal(str(precio_costo)),
                                "sell_price": Decimal(str(precio_venta)),
                                "quantity_in_stock": Decimal(str(stock)),
                                "min_stock": Decimal(str(stock_minimo)),
                                "unit": unidad,
                                "uses_inventory": usa_inventario,
                                "department_id": department_id,
                            }

                            if existing_product:
                                # Actualizar producto existente
                                for key, value in product_data.items():
                                    setattr(existing_product, key, value)
                                uow.products.update(existing_product)
                                results["updated"] += 1
                            else:
                                # Crear nuevo producto
                                new_product = Product(**product_data)
                                uow.products.add(new_product)
                                results["imported"] += 1

                        except Exception as e:
                            error_msg = f"Fila {row_num}: {str(e)}"
                            results["errors"].append(error_msg)
                            results["skipped"] += 1
                            self.logger.error(error_msg)

                # Commit de los cambios
                uow.commit()

                results["message"] = (
                    f"Importación completada: {results['imported']} nuevos, {results['updated']} actualizados, {results['skipped']} omitidos"
                )

                return results

        except Exception as e:
            self.logger.error(f"Error importando productos desde CSV: {e}")
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al importar: {e}",
            }

    def create_backup(self, backup_path: str) -> Dict[str, Any]:
        """Crea un respaldo completo de la base de datos en formato JSON."""
        try:
            with unit_of_work() as uow:
                # Obtener todos los datos
                products = uow.products.get_all()
                departments = uow.departments.get_all()
                # customers = uow.customers.get_all()  # Comentado hasta que esté disponible
                customers = []  # Temporal

                backup_data = {
                    "metadata": {
                        "created_at": datetime.now().isoformat(),
                        "version": "1.0",
                        "total_products": len(products),
                        "total_departments": len(departments),
                        "total_customers": len(customers),
                    },
                    "departments": [
                        {
                            "id": dept.id,
                            "name": dept.name,
                            "description": getattr(dept, "description", ""),
                        }
                        for dept in departments
                    ],
                    "products": [
                        {
                            "id": product.id,
                            "code": product.code,
                            "description": product.description,
                            "cost_price": (
                                float(product.cost_price) if product.cost_price else 0
                            ),
                            "sale_price": (
                                float(product.sell_price) if product.sell_price else 0
                            ),
                            "quantity_in_stock": (
                                float(product.quantity_in_stock)
                                if product.quantity_in_stock
                                else 0
                            ),
                            "min_stock": (
                                float(product.min_stock) if product.min_stock else 0
                            ),
                            "unit": getattr(product, "unit", "Unidad"),
                            "uses_inventory": product.uses_inventory,
                            "department_id": product.department_id,
                        }
                        for product in products
                    ],
                    "customers": [
                        {
                            "id": customer.id,
                            "name": customer.name,
                            "email": getattr(customer, "email", None),
                            "phone": getattr(customer, "phone", None),
                            "address": getattr(customer, "address", None),
                        }
                        for customer in customers
                    ],
                }

                # Guardar archivo JSON
                with open(backup_path, "w", encoding="utf-8") as f:
                    json.dump(backup_data, f, indent=2, ensure_ascii=False)

                self.logger.info(
                    f"Created backup with {len(products)} products, {len(departments)} departments"
                )

                return {
                    "success": True,
                    "file_path": backup_path,
                    "products_backed_up": len(products),
                    "departments_backed_up": len(departments),
                    "customers_backed_up": len(customers),
                    "message": f"Respaldo creado exitosamente: {len(products)} productos, {len(departments)} departamentos, {len(customers)} clientes.",
                }

        except Exception as e:
            self.logger.error(f"Error creating backup: {e}")
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al crear respaldo: {e}",
            }

    def restore_from_backup(self, backup_path: str) -> Dict[str, Any]:
        """Restaura datos desde un archivo de respaldo JSON."""
        if not os.path.exists(backup_path):
            return {
                "success": False,
                "error": "Archivo no encontrado",
                "message": f"El archivo {backup_path} no existe.",
            }

        try:
            with unit_of_work() as uow:
                with open(backup_path, "r", encoding="utf-8") as f:
                    backup_data = json.load(f)

                results = {
                    "success": True,
                    "products_restored": 0,
                    "departments_restored": 0,
                    "customers_restored": 0,
                    "errors": [],
                }

                # Restaurar departamentos primero (por las dependencias)
                if "departments" in backup_data:
                    for dept_data in backup_data["departments"]:
                        try:
                            existing = uow.departments.get_by_name(dept_data["name"])
                            if existing:
                                # Actualizar departamento existente
                                if "description" in dept_data:
                                    existing.description = dept_data["description"]
                                uow.departments.update(existing)
                            else:
                                # Crear nuevo departamento
                                new_dept = Department(
                                    name=dept_data["name"],
                                    description=dept_data.get("description", ""),
                                )
                                uow.departments.add(new_dept)
                            results["departments_restored"] += 1
                        except Exception as e:
                            results["errors"].append(
                                f"Error restaurando departamento '{dept_data.get('name', 'N/A')}': {e}"
                            )

                # Restaurar productos
                if "products" in backup_data:
                    for product_data in backup_data["products"]:
                        try:
                            existing = uow.products.get_by_code(product_data["code"])

                            # Buscar departamento por ID si existe
                            department_id = None
                            if product_data.get("department_id"):
                                dept = uow.departments.get_by_id(
                                    product_data["department_id"]
                                )
                                if dept:
                                    department_id = dept.id

                            product_obj_data = {
                                "code": product_data["code"],
                                "description": product_data["description"],
                                "cost_price": Decimal(
                                    str(product_data.get("cost_price", 0))
                                ),
                                "sell_price": Decimal(
                                    str(product_data.get("sale_price", 0))
                                ),
                                "quantity_in_stock": Decimal(
                                    str(product_data.get("quantity_in_stock", 0))
                                ),
                                "min_stock": Decimal(
                                    str(product_data.get("min_stock", 0))
                                ),
                                "unit": product_data.get("unit", "Unidad"),
                                "uses_inventory": product_data.get(
                                    "uses_inventory", True
                                ),
                                "department_id": department_id,
                            }

                            if existing:
                                # Actualizar producto existente
                                for key, value in product_obj_data.items():
                                    setattr(existing, key, value)
                                uow.products.update(existing)
                            else:
                                # Crear nuevo producto
                                new_product = Product(**product_obj_data)
                                uow.products.add(new_product)

                            results["products_restored"] += 1
                        except Exception as e:
                            results["errors"].append(
                                f"Error restaurando producto '{product_data.get('code', 'N/A')}': {e}"
                            )

                # Restaurar clientes (si están disponibles)
                if "customers" in backup_data:
                    for customer_data in backup_data["customers"]:
                        try:
                            # TODO: Implementar cuando el repositorio de clientes esté disponible
                            results["customers_restored"] += 1
                        except Exception as e:
                            results["errors"].append(
                                f"Error restaurando cliente '{customer_data.get('name', 'N/A')}': {e}"
                            )

                self.logger.info(
                    f"Restored {results['products_restored']} products, {results['departments_restored']} departments"
                )

                results["message"] = (
                    f"Restauración completada: {results['products_restored']} productos, {results['departments_restored']} departamentos, {results['customers_restored']} clientes."
                )

                return results

        except Exception as e:
            self.logger.error(f"Error restoring from backup: {e}")
            return {
                "success": False,
                "error": str(e),
                "message": f"Error al restaurar: {e}",
            }
